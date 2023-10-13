"""
File with MasterExcel class
"""
import os
from os import getcwd, chdir, listdir
import xlrd
import openpyxl
import csv
from copy import deepcopy
from tkinter import filedialog
from tkinter.commondialog import Dialog
from app_config.app_notices import CANCELLED, ERROR_FILE_EXTENSION, ERROR, WARNING, SUCCESS, FILE_CREATED, INFO
from tkinter.filedialog import askopenfilename, askdirectory, asksaveasfile, asksaveasfilename
from app_config.settings import DEFAULT_NAME_SAVE_FILE, EXCEL_TEMPLATE, SUPPORTED_FORMATS
from openpyxl import Workbook


class MasterExcel:

    EXPORT_DATA = deepcopy(EXCEL_TEMPLATE)
    EXPORT_DATA_NEW = []

    # def __init__(self):
    #     # self.__commands = commands
    #     self.EXPORT_DATA = deepcopy(EXCEL_TEMPLATE)

    def reset_export_data(self, reset=False):        
        self.EXPORT_DATA = deepcopy(EXCEL_TEMPLATE)
        return f'[{SUCCESS}] Data reset!'

    def _check_export_data(self):
        if self.EXPORT_DATA is None:
            return f'[{INFO}] Data undefined!'

        if len(self.EXPORT_DATA[DEFAULT_NAME_SAVE_FILE]) > 1:
            answer = input("""Data not default. 'Yes' to set to default. 'no' to add data.\nAnswer: """)
            match answer:
                case 'Yes':
                    return self.reset_export_data(True)
                case 'no':
                    return f'[{INFO}] Data be added.'
                case _:
                    return CANCELLED                

    @staticmethod
    def get_file_extension(file_name: str) -> str:
        return file_name[file_name.rfind('.'):]

    @staticmethod
    def _file_reader(file_extension, file_path) -> list:
        match file_extension:
            case '.csv':
                with open(file_path, 'r') as csvfile:
                    reader = [*csv.reader(csvfile, delimiter=';')]
            # case '.xls':


            #     reader = xlrd.open_workbook(r'C:\Users\Егор\Desktop\testxls.xls')
            #     sheet = reader.sheet_by_index(0)

            #     data = []
            #     for row in range(sheet.nrows):
            #         for cell in sheet.row(row):
            #             data.append(cell.value)
            case '.xlsx':
                worksheet = openpyxl.load_workbook(file_path).active
                reader = []
                data = []

                for i in range(0, worksheet.max_row):
                    for col in worksheet.iter_cols(1, worksheet.max_column):
                        value = col[i].value if col[i].value else ''
                        data.append(value)
                    reader.append(data.copy())
                    data.clear()                  
            case _:
                return f'[{ERROR}] Unexpected extension!'
        return reader

    @staticmethod
    def left_titles():
        answer = input("""Left the titles of columns? Yes to accept, no to cut.\nAnswer: """).strip()
        match answer:
            case 'Yes':
                return True
            case 'no':
                return False
            case _:
                return None
