"""
!!! - IN PROGRESS - !!!
"""
import openpyxl


class Checker:

    def __init__(self):
        self.__path = r"C:\py_proj\ExcelMaster\test_folder\sample2.xlsx"
    
    def check(self):
        worksheet = openpyxl.load_workbook(self.__path).active
        reader = []
        already_in = []

        for i in range(0, worksheet.max_row):

            data = [col[i].value if col[i].value else '' for col in worksheet.iter_cols(1, worksheet.max_column)]

            if data[-1] not in already_in and self.single_true(data):
                already_in.append(data[-1])
                reader.append(data.copy())
            data.clear()

        print(already_in)
        print(reader)

    @staticmethod
    def single_true(iterable):
        i = iter(iterable)
        return any(i) and not any(i)


checker = Checker()

checker.check()

print(checker.single_true(['Test', '', '', '']))
