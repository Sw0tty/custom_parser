import openpyxl
import datetime
import os
import csv
from tkinter import filedialog
import pandas
import requests
from bs4 import BeautifulSoup as bs


SUPPORTED_FORMATS = [('CSV file', '*.csv'), ('Excel book', '*.xlsx'), ]

OPENPX_EXCEL_TEMPLATE = [
    ['Закупки по', 'Наименование закупки', 'Начальная (максимальная) цена контракта',
         'Наименование заказчика', 'Дата окончания подачи заявок', 'Интерес',
         'Категория', 'Ссылка']
         ]

os.chdir(os.getcwd() + r'\test_folder')

wb = openpyxl.Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = "TEST"

# # Rows can also be appended
# ws.append([1, 2, 3])

# # Python types will automatically be converted
# ws['A2'] = datetime.datetime.now()

# # Rename active sheet
# ws.title = 'TEST'

# Save the file


# sheet = pandas.read_excel(r'C:\Users\Егор\Desktop\testxls.xls')

class A:

    def __init__(self, vla) -> None:
        self.a_param = vla
    
    @property
    def a_param(self):
        return self.a
    
    @a_param.setter
    def a_param(self, new):
        self.a = new

    def fff(self, new):
        if new == 1:
            return 1
        

a = A(1)

# print(a.a_param)

# a.a_param = 45

# print(a.a_param)

print(bool(a.fff(8)))


url = 'https://zakupki.gov.ru/epz/order/extendedsearch/results.html?morphology=on&search-filter=Дате+размещения&pageNumber=1&sortDirection=false&recordsPerPage=_10&showLotsInfoHidden=false&sortBy=UPDATE_DATE&fz44=on&fz223=on&af=on&priceFromGeneral=5000000&currencyIdGeneral=-1&publishDateFrom=03.10.2023&publishDateTo=05.10.2023'

response = requests.get(url, headers= {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'
})
html = response.text
 
soup = bs(html, 'html.parser')

print(type(soup))

if isinstance(soup, bs):
    print(1)

pagination = soup.find('div', class_='paginator-block')
pages = pagination.find_all('span', class_='link-text')

if pages:


    print(pages[-1].text)

# for page in pages:

#     print(page)

# try:
#     file = filedialog.asksaveasfilename(initialfile='name', title="Save file", initialdir=os.getcwd(), filetypes=(*SUPPORTED_FORMATS, ('All files', '*')), defaultextension=True)
#     if not file:
#         print(12313)
# except FileNotFoundError:   
#     print('123')


# extansion = file[file.rfind('.'):]

# match extansion:
#     case '.xlsx':
#         wb.save(file)

#     case '.csv':
#         with open(file, 'w') as csvfilewrite:
#             writer = csv.writer(csvfilewrite, lineterminator="\r", delimiter = ";")
            
#             for i in OPENPX_EXCEL_TEMPLATE:
#                 writer.writerow(i)




# path_dir = filedialog.asksaveasfile(filetypes=(*SUPPORTED_FORMATS, ('All files', '*')))

# path_dir = filedialog.askopenfilename(filetypes=(*SUPPORTED_FORMATS, ('All files', '*')))

# print(path_dir)

# worksheet = openpyxl.load_workbook(os.getcwd() + "\sample2.xlsx").active
# reader = []
# data = []

# for i in range(0, worksheet.max_row):
#     count = 0
#     for col in worksheet.iter_cols(1, worksheet.max_column):
#         value = col[i].value if col[i].value else ''
#         if count == worksheet.max_column - 1:
#             print(col[i].value)
#         data.append(value)
#         count += 1
#     reader.append(data.copy())
#     data.clear()   

# print(reader)
